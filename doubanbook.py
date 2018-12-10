'''
	遍历所有页面，通过next href判断是否结束	
'''
from bs4 import BeautifulSoup
import time
import requests
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import os

agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.110 Safari/537.36'
headers = {
	'User-Agent':agent,
	'Host':'book.douban.com',
	'Referer':'https://book.douban.com',
	'Connection':'keep-alive',
}
TYPE = 'T'	#T：综合排序；S：按评价排序；R:按出版日期排序
filename = 'doubanbook.xlsx'
target = 'https://book.douban.com'
tag_url = 'https://book.douban.com/tag/?view=type'

def GetTags():
	tag_list = []
	html = requests.get(tag_url, headers = headers).text
	soup = BeautifulSoup(html, 'lxml')
	soup_list = soup.findAll('table', {'class':'tagCol'})
	for tags in soup_list:
		tags = tags.findAll('a')
		for tag in tags:
			tag = tag.string.strip()
			tag_list.append(tag)
	return tag_list

#获取标签下的首个页面url
def GetUrl(tag):
	url = target + '/tag/' + tag + '?start=0&type=' + TYPE
	return url

def GetContent(url):
	book_list = []
	while True:
		html = requests.get(url, headers = headers).text
		soup = BeautifulSoup(html, 'lxml')
		soup_list = soup.find('ul', {'class':'subject-list'})
		time.sleep(np.random.rand()*10)

		for book_info in soup_list.findAll('div', {'class':'info'}):
			pub_list = book_info.find('div', {'class':'pub'}).string.strip().split('/')
			book_url = book_info.find('a').get('href')
			try:
				title = book_info.find('a').string.strip()
			except:
				title0 = book_info.find('a').get('title').strip()
				title1 = book_info.find('a').find('span').string.strip()
				title = title0 + title1
			try:
				author = '/'.join(pub_list[0:-3])
			except:
				author = '暂无'
			try:
				pub = '/'.join(pub_list[-3:])
			except:
				pub = '暂无'
			try:
				rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
			except:
				rating = '0.0'
			try:
				people_num = book_info.find('span', {'class':'pl'}).string.strip()[1:-4]
				people_num = int(people_num)
			except:
				people_num = '0'
			book_list.append([title, rating, people_num, author, pub, book_url])
		try:
			href = soup.find('span', {'class':'next'}).find('a').get('href')
			url = target + href
		except:
			break
	book_list = sorted(book_list, key = lambda x:x[1],reverse = True)
	return book_list

def save_xlsx(tag, book_list):
	if not os.path.exists(filename):
		wb = Workbook()
		wb.save(filename)
	wb = load_workbook(filename)
	ws = wb.create_sheet(tag, 0)
	ws.append(['序号', '书名', '评分', '评价人数', '作者', '出版信息', '链接'])
	count = 1
	for book in book_list:
		ws.append([count, book[0], float(book[1]), int(book[2]), book[3], book[4], book[5]])
		count += 1
	wb.save(filename)

def main():
	tag_list = GetTags()
	print('所有标签:\n')
	print(tag_list)

#遍历所有标签
	for tag in tag_list:
		print('将保存标签:"{}"\n'.format(tag))
		url = GetUrl(tag)
		book_list = GetContent(url)
		save_xlsx(tag, book_list)
		print('完成！\n')
	#tag = input('输入需要保存的标签:')
	#print('将保存标签:"{}"\n'.format(tag))
	#url = GetUrl(tag)
	#book_list = GetContent(url)
	#save_xlsx(tag, book_list)
	#print('完成！\n')

if __name__ == '__main__':
	main()
